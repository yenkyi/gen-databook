import netmiko
from netmiko import ConnectHandler, NetmikoAuthenticationException, NetmikoTimeoutException, ConfigInvalidException
from concurrent.futures import ThreadPoolExecutor
from pprint import pprint
from itertools import repeat
import logging
import time
import getpass
import re
import csv
import os
import os.path
import sys
import textfsm
import openpyxl
from datetime import date
from datetime import time
from datetime import datetime
from openpyxl.styles import Font, Fill,Alignment, Border, Side, DEFAULT_FONT, PatternFill
from openpyxl import load_workbook
from operator import itemgetter



def read_map(device_list_file):
    
    map_data = {}

    with open(device_list_file) as csv_datafile:
        csv_reader = csv.reader(csv_datafile, delimiter=';')
        
        for row in csv_reader:
            if row :
                map_data[row[0]] = [row[0],row[1],row[2],row[3],row[4]]

    return map_data           

def gen_databook_table(databook_input_file,databook_output_file,meta_info):
    # Load the input file to a variable
    input_file = open(databook_input_file, encoding='utf-8')
    raw_text_data = input_file.read()
    input_file.close()
    
    #Remove-Inputfile
    if os.path.exists(databook_input_file):
        os.remove(databook_input_file)
    
    # Run the text through the FSM. 
    # The argument 'template' is a file handle and 'raw_text_data' is a 
    # string with the content from the show_inventory.txt file
    template = open("fsm/databook_multiple.textfsm")
    re_table = textfsm.TextFSM(template)
    fsm_results = re_table.ParseText(raw_text_data)
    
    #print(fsm_results)
      
    #Writing to excel file
    #check If Workbook File Exists
    if os.path.exists(databook_output_file):
        wb = load_workbook(filename = databook_output_file)
    else:
    #create a new Workbook
        wb = openpyxl.Workbook()
    
    if meta_info["hostname"] in wb.sheetnames:
        wb.remove(wb[meta_info["hostname"]])
    
    sheet = wb.create_sheet(title=meta_info["hostname"])
    
    sheet.cell(row=4,column=1).value = "Hostname : " 
    sheet.cell(row=4,column=2).value = meta_info["hostname"]
    sheet.cell(row=4,column=2).font = Font(color='000000FF')
    
    #IP Address 
    sheet.cell(row=5,column=1).value = "IP Address : " 
    sheet.cell(row=5,column=2).value = meta_info["ip_address"]
    sheet.cell(row=5,column=2).font = Font(color= '000000FF')
    
    #Location
    sheet.cell(row=6,column=1).value = "Location : " 
    sheet.cell(row=6,column=2).value = meta_info["location"]
    sheet.cell(row=6,column=2).font = Font(color= '000000FF')
    
    sheet.cell(row=4,column=5).value =  meta_info["serial_number"]
    
    #Bold Hostname
    sheet.merge_cells('A1:J1')
    sheet.cell(row=1,column=1).value = meta_info["hostname"]
    sheet.cell(row=1,column=1).font = Font(bold=True,size=16)
    sheet.cell(row=1,column=1).alignment = Alignment(horizontal="center", vertical="center")
    for colz in range (1,13):
        sheet.cell(row=1, column = colz).fill = PatternFill('solid', fgColor = '57b830')
        sheet.cell(row=1, column = colz).border = Border(bottom=Side(border_style='medium', color='FF000000'),
                                          right=Side(border_style='medium', color='FF000000'))
    
    
    #Serial Formating
    sheet.merge_cells('E4:G6') 
    sheet.cell(row=4,column=5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    #Space between The serial numbers
    sheet.cell(row=10,column=1).value = ""
    
    
    # set the width of the column 
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['E'].width = 24
    sheet.column_dimensions['H'].width = 16
    sheet.column_dimensions['I'].width = 16
    sheet.column_dimensions['J'].width = 20
    
    #Headers
    sheet.append(re_table.header)
    header_row = 11
    sheet.cell(row=header_row,column=1).value = "PORT"
    sheet.cell(row=header_row,column=2).value = "DESCRIPTION"
    sheet.cell(row=header_row,column=3).value = "STATUS"
    sheet.cell(row=header_row,column=4).value = "DUPLEX"
    sheet.cell(row=header_row,column=9).value = "NE ID"
    sheet.cell(row=header_row,column=12).value = "PORT"
    
    
    for row in fsm_results:
        sheet.append(row)
    
    last_row_num = len(fsm_results)+14
    
    for roz in range(11,last_row_num):
        sheet.cell(row=roz, column = 8).border = Border(left=Side(border_style='medium',color='FF000000'))
    
    #Top Header Lines
    sheet.row_dimensions[11].height = 16
    for col in range(1,13):
        sheet.cell(row=11, column = col).font = Font(bold=True)
        sheet.cell(row=11, column = col).fill = PatternFill('solid', fgColor = '0099CCFF')
        sheet.cell(row=11, column = col).border = Border(top=Side(border_style='medium',
                         color='FF000000'),
                bottom=Side(border_style='thick',
                            color='FF000000')
				)
                
    #Beginning of the NEIGHBOR            
    sheet.cell(row=11, column = 8).border = Border(left=Side(border_style='medium',color='FF000000'),top=Side(border_style='medium',
                         color='FF000000'),bottom=Side(border_style='thick',color='FF000000'))
    
    sheet.cell(row=10, column = 8).value = "Connected To"
    sheet.merge_cells('H10:J10')
    sheet.cell(row=10, column = 8).font = Font(bold=True, size=14)
    sheet.cell(row=10, column = 8).fill = PatternFill('solid', fgColor = '57b830')
    sheet.cell(row=10, column = 8).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=10, column = 8).border = Border(top=Side(border_style='medium', color='FF000000'),left=Side(border_style='medium', color='FF000000'),
                                          right=Side(border_style='medium', color='FF000000'),
				)
    for colz in range (9,13):
        sheet.cell(row=10, column = colz).fill = PatternFill('solid', fgColor = '57b830')
        sheet.cell(row=10, column = colz).border = Border(top=Side(border_style='medium', color='FF000000'), 
                    right=Side(border_style='medium', color='FF000000'))
    
    #Align the Headers to Center
    for colz in range (1,13):
        sheet.cell(row=11, column = colz).alignment = Alignment(horizontal="center", vertical="center")
    
    
    #Bottom Line
    for col in range(1,13):
        sheet.cell(row=last_row_num, column = col).border = Border(top=Side(border_style='medium',
                         color='FF000000')
				)
    
    for rowz in range(2,last_row_num):
        sheet.cell(row=rowz, column = 13).border = Border(left=Side(border_style='medium',
                         color='FF000000')
				)
    #print VLANS on TRUNKS 
    sheet.cell(row=last_row_num+2,column=1).value =  "VLANs on TRUNK PORTS"
    sheet.cell(row=last_row_num+2, column = 1).font = Font(bold=True, size=12)
    sheet.cell(row=last_row_num+3,column=1).value =  meta_info["trunk_vlans"]
    sheet.cell(row=last_row_num+3,column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    #Merge the VLANs List
    trunk_start_row=last_row_num+3
    trunk_end_row = last_row_num+14
    cell_to_merge = str("A"+str(trunk_start_row)+":"+"J"+str(trunk_end_row))
    #print(cell_to_merge)
    
    #Bottom Line
    for col in range(1,13):
        sheet.cell(row=trunk_end_row+2, column = col).border = Border(top=Side(border_style='medium',
                         color='FF000000')
				)
    for rowz in range(2,trunk_end_row+2):
        sheet.cell(row=rowz, column = 13).border = Border(left=Side(border_style='medium',
                         color='FF000000')
				)
    
    sheet.merge_cells(cell_to_merge) 
    #for x in range(last_row_num+3, last_row_num+7):
    #    sheet.merge_cells(start_row=x, start_column=1, end_row=x, end_column=12)
    
    sheet.delete_cols(10,2)
    
    #Set Default FONT and Save Workbook
    DEFAULT_FONT.name = "Arial"
    wb.save(databook_output_file)
    
    
    
    #outfile.close()
    template.close()
    return 0
    

def send_config_command(device_dict):
    start_msg = '===> {} Connection: {}'
    received_msg = '<=== {} Received:   {}'
    ip = device_dict["ip"]+" - "+device_dict["hostname"]
    #ip = device_dict["ip"]
            
    logging.info(start_msg.format(datetime.now().time(), ip))
    ##if ip == '192.168.100.1': time.sleep(5)
    
    
    ###NEW PART
    
    network_node  = {'device_type':'cisco_ios', 
                    'ip':device_dict['ip'],
                    'username' : device_dict['username'],
                    'password' : device_dict['password'],
                    'secret' : device_dict['secret'],
                    }
       
    dev_locate = device_dict['site']
    host_ip = device_dict['ip']
    host_name = device_dict['hostname']
    folder_name_cdp = device_dict['folder_dbook']
            
    try:
        with ConnectHandler(**network_node) as ssh:
            ssh.enable()        
                                   
            output_term_mon = ssh.send_command('terminal length 0')
            
            int_status_output = ssh.send_command('show int status')
            
            #print(int_status_output)
            
            serial_num = ssh.send_command('show version | in System Serial|System serial')
            gebauede = ssh.send_command('show snmp location')
            
            output_vlans = ssh.send_command('show int trunk | beg allowed on trunk')
            
            #pprint(output_vlans)
            
            get_vlans = re.compile(r'(?s) on trunk(.*?)(?:(?:\r*\n){2})')
            sw_vlans = get_vlans.findall(output_vlans)
            
            
            get_interfaces = re.compile(r'([TeGi]{2}\d.\d.\d+).*?')
            #get_interfaces = re.compile(r'([TeGi]{2}\d.\d+).*?')
            interfaces_list = get_interfaces.findall(int_status_output)
            
            #print(interfaces_list)
            
            file_dir_cdp = str(folder_name_cdp)
            
                         
            if not os.path.exists(file_dir_cdp):
                os.makedirs(file_dir_cdp)
                        
            for i in interfaces_list:
                cdp_output = ssh.send_command("show cdp nei "+i+" | inc Ten|Gig")
                int_status_output = ssh.send_command("show int "+i+" status | inc Te|Gi")
                
                #print(inventory_output)
                
                f_cdp = open(file_dir_cdp+"/"+host_name+"_DBOOK-INTF.txt","a+")
                f_cdp.write(int_status_output)
                f_cdp.write("  ")
                f_cdp.write(cdp_output)
                f_cdp.write("\n\n--------------------------------------------------\n\n")
                f_cdp.close()
            
            #Generate the databook for the switch
            meta_data = {
            'hostname' : host_name,
            'ip_address' : host_ip,
            'serial_number': serial_num,
            'location' : gebauede,
            'trunk_vlans': sw_vlans[0],
            'site' : dev_locate
            }
                      
            logging.info(received_msg.format(datetime.now().time(), ip))

            return meta_data
    except Exception as err:        
    #except (NetmikoTimeoutException, NetmikoAuthenticationException, ConfigInvalidException) as err:
        logging.warning(err)


def send_command_to_devices(devices):
    data = {}
    with ThreadPoolExecutor(max_workers=10) as executor:
        result = executor.map(send_config_command, devices)
        return result
        
        
if __name__ == "__main__":

    time_now = datetime.now()
    dt_string = time_now.strftime("%d-%m-%Y_%H")

    try:
        dev_list_file = sys.argv[1]
    except:
        raise SystemExit(f"Usage: {sys.argv[0]} input-list.csv")

    logging.getLogger('paramiko').setLevel(logging.WARNING)

    logging.basicConfig(
        format = '%(threadName)s %(name)s %(levelname)s: %(message)s',
        level=logging.INFO)
    
    folder_name_cdp = "DBOOK_STATUS_"+dt_string+"Uhr"
    #folder_name = "INVENTORY-"+str(date.today())+"-"+str(time.hour)
    #print("Folder_name_cdp"+folder_name_cdp+"\n")
    
    map_return_values = read_map(dev_list_file)
    
    #print(map_return_values)
    
    
    device_list = []
    site_list = []
    
    for device_name in map_return_values:
        #print(map_return_values.get(device_name)[1])
        host_name = map_return_values.get(device_name)[0]
        host_ip   = map_return_values.get(device_name)[1]
        dev_locate= map_return_values.get(device_name)[2]
        dev_username=map_return_values.get(device_name)[3]
        dev_pass= map_return_values.get(device_name)[4]
        
        network_node  = {'device_type':'cisco_ios', 
                        'ip':host_ip,
                        'username' : dev_username,
                        'password' : dev_pass,
                        'secret' : dev_pass,
                        'site' : dev_locate,
                        'folder_dbook' : folder_name_cdp,
                        'hostname' : host_name
                        }
        device_list.append(network_node)
        site_list.append(dev_locate)
    
    output_from_device_unsorted = send_command_to_devices(device_list)
    
    #Convert Generator Object to List
    output_from_device_list = []
    for return_device in output_from_device_unsorted:
        #pprint(return_device)
        if return_device != None:
            output_from_device_list.append(return_device)
            
    output_from_device = sorted(output_from_device_list, key=itemgetter('hostname'))
    #pprint(output_from_device)
    
    info_msg = '== {} Generating Excel : {}'
    
    for device_ip in output_from_device:
        if os.path.exists(folder_name_cdp+"/"+device_ip['hostname']+"_DBOOK-INTF.txt"):
            logging.info(info_msg.format(datetime.now().time(), device_ip['hostname']))
            input_file_name = folder_name_cdp+"/"+device_ip['hostname']+"_DBOOK-INTF.txt"
            output_file_name = folder_name_cdp+"/"+device_ip['site']+"_DBOOK.xlsx"
            gen_databook_table(input_file_name,output_file_name,device_ip)
       
    
    
    